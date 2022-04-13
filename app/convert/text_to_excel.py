from logging import getLogger
from pathlib import Path
from openpyxl.cell import Cell
from typing import Any, Callable, Dict, Iterator, Tuple

from config.config_class import ConfigApp
from dataobject.data_object import DataObject
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from util.resource_path import resource_path
from openpyxl.cell import Cell

from .check_version import check_datafile_version
from .class_definitions import Definitions
from .converter import Converter
from .load_definitions import load_definisions
from .set_metadata import set_metadata
from .write_keyvalue import write_keyvalue
from .write_table import write_table

applogger = getLogger('app')
filelogger = getLogger('file')


class TextToExcel(Converter):
    def __init__(self, config: ConfigApp) -> None:
        super().__init__(config)

        self.read_action: Dict[str, Callable] = {
            'key-value': write_keyvalue,
            'table': write_table
        }

    def read(self, datadict: Dict[str, Any], *, filename: Path) -> None:
        applogger.info('@read')
        self.data.update(datadict)
        current, latest = check_datafile_version(datadict.get('$metadata', {}), self.config.options.definitions)
        applogger.info('@checked workbook version')
        applogger.info('CurrentVersion: %s', current)
        applogger.info('LatestVersion: %s', latest)
        if not current:
            raise Exception
        self.definition_data = load_definisions(current)
        filelogger.debug(self.definition_data)
        self.metadata = set_metadata(self.data, self.definition_data, filename, 'excel')
        template_path = resource_path(self.config.options.templates.folder, self.definition_data.template_name)
        if template_path.exists():
            self.workbook = load_workbook(template_path.as_posix())
        else:
            self.workbook = Workbook()

        for name in self.definition_data.includes:
            applogger.info(f'Processing: {name}')
            definitions = self.definition_data.definitions.get(name, None)
            if not definitions or not self.read_action.get(definitions.type) or self.workbook[definitions.sheet] is None:
                continue
            action: Callable[[Workbook, DataObject, Definitions], Iterator[Tuple[Cell, Any]]
                             ] = self.read_action.get(definitions.type, lambda: print('Unknown function'))
            for c, v in action(self.workbook, self.data, definitions):
                c.value = v

    def write(self, path: Path) -> None:
        applogger.info('@write')
        self.workbook.save(path.as_posix())
